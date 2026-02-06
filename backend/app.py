from flask import Flask, jsonify, request, send_file, Response
from flask_cors import CORS
import os, sys, json, base64, time, threading, sqlite3
from datetime import datetime, timedelta
from typing import List, Dict, Any
from cryptography.hazmat.primitives.asymmetric import ec
from cryptography.hazmat.primitives import serialization

try:
  from openpyxl import load_workbook
except Exception:
  load_workbook = None

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import getBanBiao as gb
from menu_system.api import bp as menu_bp
from menu_system.services import init_db as menu_init_db

app = Flask(__name__)
CORS(app)
menu_init_db()
app.register_blueprint(menu_bp)

@app.get('/api/v1/borrow/stats')
def borrow_stats():
  date = request.args.get('date')
  data = gb.get_banbiao_data(date)
  return jsonify(data)

@app.post('/api/v1/products')
def products():
  payload_in = request.json or {}
  seldate = int(payload_in.get('seldate', 0))
  chooseData = payload_in.get('chooseData')
  custom_start_date = payload_in.get('custom_start_date')
  custom_end_date = payload_in.get('custom_end_date')
  data = gb._fetch_products(custom_start_date, custom_end_date, seldate, chooseData)
  return jsonify({'data': data})

# ===== Sorting and upload integration =====
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'data')

def _ensure_dir(p):
  try:
    os.makedirs(p, exist_ok=True)
  except Exception:
    pass

# ===== Public image serving =====
IMAGES_DIR = os.path.join(UPLOAD_DIR, 'images')
_ensure_dir(IMAGES_DIR)

@app.get('/uploads/images/<fname>')
def serve_uploaded_image(fname):
  try:
    import re
    if not re.fullmatch(r'[A-Za-z0-9_.-]+', fname or ''):
      return jsonify({'error': 'invalid filename'}), 400
    p = os.path.join(IMAGES_DIR, fname)
    if not os.path.isfile(p):
      return jsonify({'error': '文件不存在'}), 404
    ext = os.path.splitext(fname)[1].lower()
    mt = 'image/jpeg'
    if ext == '.png':
      mt = 'image/png'
    elif ext in ['.jpg', '.jpeg']:
      mt = 'image/jpeg'
    else:
      mt = 'application/octet-stream'
    return send_file(p, mimetype=mt)
  except Exception as e:
    return jsonify({'error': str(e)}), 500

# ===== Ledger file store persistence =====
LEDGER_DIR = os.path.join(UPLOAD_DIR, 'ledger')
LEDGER_STORE = os.path.join(LEDGER_DIR, 'ledger_store.json')

def _read_ledger_store():
  try:
    _ensure_dir(LEDGER_DIR)
    if not os.path.exists(LEDGER_STORE):
      return {}
    import json
    with open(LEDGER_STORE, 'r', encoding='utf-8') as f:
      return json.load(f)
  except Exception:
    return {}

def _write_ledger_store(data: dict):
  try:
    _ensure_dir(LEDGER_DIR)
    import json
    with open(LEDGER_STORE, 'w', encoding='utf-8') as f:
      json.dump(data, f, ensure_ascii=False, indent=2)
    return True
  except Exception:
    return False

def _ledger_active_paths():
  st = _read_ledger_store()
  at = (st.get('active_template') or '').strip()
  ac = (st.get('active_conversion') or '').strip()
  return at if at else None, ac if ac else None

def _sort_order_path():
  _ensure_dir(UPLOAD_DIR)
  return os.path.join(UPLOAD_DIR, 'sort_order.txt')

def _read_text_file(path):
  try:
    with open(path, 'rb') as f:
      data = f.read()
    for enc in ('utf-8', 'utf-8-sig', 'utf-16', 'utf-16-le', 'utf-16-be', 'gb18030'):
      try:
        return data.decode(enc)
      except Exception:
        continue
    return data.decode('utf-8', errors='ignore')
  except Exception as e:
    raise e

@app.get('/api/v1/sort-order')
def get_sort_order():
  try:
    p = _sort_order_path()
    if not os.path.exists(p):
      return jsonify({'order': []})
    content = _read_text_file(p)
    lines = [x.strip() for x in content.splitlines() if x.strip()]
    return jsonify({'order': lines})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.post('/api/v1/sort-order')
def set_sort_order():
  try:
    body = request.json or {}
    text = (body.get('text') or '').strip()
    order = body.get('order')
    lines = []
    if isinstance(order, list):
      lines = [str(x).strip() for x in order if str(x).strip()]
    elif text:
      lines = [x.strip() for x in text.splitlines() if x.strip()]
    p = _sort_order_path()
    _ensure_dir(os.path.dirname(p))
    with open(p, 'w', encoding='utf-8') as f:
      f.write('\n'.join(lines))
    return jsonify({'saved': True, 'count': len(lines)})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.post('/api/v1/upload/sort-file')
def upload_sort_file():
  try:
    if 'file' not in request.files:
      return jsonify({'error': '缺少文件'}), 400
    file = request.files['file']
    data = file.read()
    # 兼容多种文本编码，避免上传的文件为 UTF-16/GB18030 导致内容异常
    for enc in ('utf-8', 'utf-8-sig', 'utf-16', 'utf-16-le', 'utf-16-be', 'gb18030'):
      try:
        content = data.decode(enc)
        break
      except Exception:
        content = None
    if content is None:
      content = data.decode('utf-8', errors='ignore')
    lines = [x.strip() for x in content.splitlines() if x.strip()]
    p = _sort_order_path()
    _ensure_dir(os.path.dirname(p))
    with open(p, 'w', encoding='utf-8') as f:
      f.write('\n'.join(lines))
    return jsonify({'saved': True, 'count': len(lines)})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.post('/api/v1/upload/excel')
def upload_excel():
  try:
    if 'file' not in request.files:
      return jsonify({'error': '缺少文件'}), 400
    file = request.files['file']
    fname = file.filename or 'import.xlsx'
    _ensure_dir(UPLOAD_DIR)
    save_path = os.path.join(UPLOAD_DIR, fname)
    file.save(save_path)
    return jsonify({'saved': True, 'path': save_path})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

# ===== Order data persistence (file-based) =====
ORDER_STORE = os.path.join(UPLOAD_DIR, 'order.json')

def _read_order_store():
  try:
    if not os.path.exists(ORDER_STORE):
      return {}
    with open(ORDER_STORE, 'r', encoding='utf-8') as f:
      import json
      return json.load(f)
  except Exception:
    return {}

def _write_order_store(data: dict):
  try:
    _ensure_dir(UPLOAD_DIR)
    import json
    with open(ORDER_STORE, 'w', encoding='utf-8') as f:
      json.dump(data, f, ensure_ascii=False, indent=2)
    return True
  except Exception:
    return False

@app.get('/api/v1/order/items')
def order_items_get():
  try:
    store = _read_order_store()
    date = (request.args.get('date') or '').strip()
    items_by_date = store.get('items_by_date') or {}
    items = []
    if date:
      items = items_by_date.get(date) or []
    else:
      items = store.get('items') or []

    # Fallback: try to load from legacy order data under 
    # <repo_root>/订货/data/*.json if current store is empty
    if not items:
      legacy = _load_legacy_order_items()
      if legacy:
        items = legacy
        from datetime import datetime
        store['items'] = items
        store['lastModified'] = datetime.utcnow().isoformat()
        _write_order_store(store)

    # Enrich items from legacy when unit/limit missing
    if items:
      enriched = _enrich_items_with_legacy(items)
      if enriched:
        items = enriched
        from datetime import datetime
        store['items'] = items
        store['lastModified'] = datetime.utcnow().isoformat()
        _write_order_store(store)
    return jsonify({'items': items, 'lastModified': store.get('lastModified'), 'date': date or None})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.post('/api/v1/order/items')
def order_items_set():
  try:
    body = request.json or {}
    items = body.get('items')
    if not isinstance(items, list):
      return jsonify({'error': 'items 必须为数组'}), 400
    date = (request.args.get('date') or '').strip()
    store = _read_order_store()
    if date:
      items_by_date = store.get('items_by_date') or {}
      items_by_date[date] = items
      store['items_by_date'] = items_by_date
    else:
      store['items'] = items
    from datetime import datetime
    store['lastModified'] = datetime.utcnow().isoformat()
    ok = _write_order_store(store)
    return jsonify({'saved': ok, 'count': len(items), 'date': date or None})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.post('/api/v1/order/save')
def order_save_key():
  try:
    body = request.json or {}
    key = (body.get('key') or '').strip()
    data = body.get('data')
    if not key:
      return jsonify({'error': '缺少 key'}), 400
    store = _read_order_store()
    store[key] = data
    from datetime import datetime
    store['lastModified'] = datetime.utcnow().isoformat()
    ok = _write_order_store(store)
    return jsonify({'saved': ok})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

def _load_legacy_order_items():
  try:
    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    legacy_dir = os.path.join(root_dir, 'order', 'data')
    if not os.path.isdir(legacy_dir):
      return []
    import json
    from glob import glob
    files = glob(os.path.join(legacy_dir, '*.json'))
    out = []
    for fp in files:
      try:
        with open(fp, 'r', encoding='utf-8') as f:
          data = json.load(f)
        # find latest history_* list entry by timestamp if present
        latest = None
        latest_ts = ''
        for k, v in data.items():
          if not k.startswith('history_'):
            continue
          if isinstance(v, list) and v:
            # prefer last item or max timestamp
            cand = v[-1]
            ts = ''
            if isinstance(cand, dict):
              ts = str(cand.get('timestamp') or '')
            if ts > latest_ts:
              latest = cand
              latest_ts = ts
        if isinstance(latest, dict):
          arr = latest.get('originalData') or []
          if isinstance(arr, list):
            for it in arr:
              try:
                m = it if isinstance(it, dict) else {}
                pid = str(m.get('code') or m.get('id') or '').strip()
                name = str(m.get('name') or '').strip()
                spec = str(m.get('spec') or '').strip()
                qty = str(m.get('qtyVal') or m.get('qty') or '').strip()
                if not pid and not name:
                  continue
                out.append({
                  'product_id': pid,
                  'product_name': name,
                  'spec': spec,
                  'quantity': qty,
                  # preserve legacy fields for unit/limit and conversion
                  'unit': str(m.get('point') or ''),
                  'limit': str(m.get('dian') or ''),
                  'upb': m.get('upb'),
                  'isBoxUnit': m.get('isBoxUnit'),
                  'qtyVal': str(m.get('qtyVal') or ''),
                  'initBox': str(m.get('initBox') or ''),
                })
              except Exception:
                continue
      except Exception:
        continue
    return out
  except Exception:
    return []

def _enrich_items_with_legacy(items):
  try:
    legacy = _load_legacy_order_items()
    if not legacy:
      return None
    def _norm_name(s):
      return str(s or '').replace('\s', '')
    by_id = {str(x.get('product_id') or ''): x for x in legacy if str(x.get('product_id') or '')}
    by_name = { _norm_name(x.get('product_name')): x for x in legacy if _norm_name(x.get('product_name')) }
    changed = False
    for it in items:
      unit = (it.get('unit') or '').strip()
      limit = (it.get('limit') or '').strip()
      upb = it.get('upb')
      is_box = it.get('isBoxUnit')
      if unit and limit and upb is not None and is_box is not None:
        continue
      src = None
      pid = str(it.get('product_id') or '')
      nm = _norm_name(it.get('product_name'))
      if pid and pid in by_id:
        src = by_id[pid]
      elif nm and nm in by_name:
        src = by_name[nm]
      if not src:
        continue
      if not unit:
        it['unit'] = src.get('unit') or ''
      if not limit:
        it['limit'] = src.get('limit') or ''
      if upb is None:
        it['upb'] = src.get('upb')
      if is_box is None:
        it['isBoxUnit'] = src.get('isBoxUnit')
      if not it.get('qtyVal'):
        it['qtyVal'] = src.get('qtyVal') or ''
      if not it.get('initBox'):
        it['initBox'] = src.get('initBox') or ''
      changed = True
    return items if changed else None
  except Exception:
    return None

def _read_conversion_table(conv_path):
  if load_workbook is None:
    raise RuntimeError('openpyxl not available')
  wb = load_workbook(conv_path, data_only=True)
  sh = wb.worksheets[0]
  headers = {}
  for row in sh.iter_rows(min_row=1, max_row=5):
    for c in row:
      val = str(c.value or '').strip()
      if not val:
        continue
      headers[val] = c.column
  def _find(cols):
    for name, col in headers.items():
      for k in cols:
        if k in name:
          return col
    return None
  col_code = _find(['商品编码', '编码', '商品编号', '代码', '货品编码'])
  col_unitkg = _find(['进货单位千克', '单位千克', 'kg', '千克'])
  if not col_code or not col_unitkg:
    raise RuntimeError('无法定位转换表列')
  m = {}
  for row in sh.iter_rows(min_row=2):
    code = row[col_code-1].value
    unitkg = row[col_unitkg-1].value
    if code is None:
      continue
    s = str(code).strip()
    if not s:
      continue
    try:
      uk = float(str(unitkg).strip()) if unitkg is not None else 0.0
    except Exception:
      uk = 0.0
    m[s] = uk
  return m

def _locate_headers(sh):
  hdr_row = None
  for r in range(1, min(15, sh.max_row+1)):
    names = [str(sh.cell(row=r, column=c).value or '').strip() for c in range(1, min(40, sh.max_column+1))]
    if any(('编码' in x or '编号' in x or '代码' in x) for x in names):
      hdr_row = r
      break
  if hdr_row is None:
    hdr_row = 1
  headers = {}
  for c in range(1, sh.max_column+1):
    name = str(sh.cell(row=hdr_row, column=c).value or '').strip()
    if name:
      headers[name] = c
  def _col(*keys):
    for k in keys:
      for name, col in headers.items():
        if k in name:
          return col
    return None
  col_code = _col('产品代码', '产品编号', '产品编码', '商品编码', '商品编号', '商品代码', '编码', '货品编码', '代码')
  col_qty = _col('采购数量', '进货数量', '数量')
  col_in_date = _col('进货日期', '采购日期')
  col_prod_date = _col('生产日期')
  col_kg = _col('千克', '公斤', 'KG')
  return hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg

def _as_bool(v, default=None):
  if v is None:
    return default
  s = str(v).strip().lower()
  if s in ['1','true','yes','y','on']:
    return True
  if s in ['0','false','no','n','off']:
    return False
  return default

def _sanitize_url(u):
  return str(u or '').replace('`','').strip()

@app.post('/api/v1/ledger/external-upload')
def ledger_external_upload():
  try:
    body = request.json or {}
    form = request.form or {}
    path = (body.get('path') or form.get('path') or request.args.get('path') or '').strip()
    token = (body.get('token') or form.get('token') or request.args.get('token') or os.environ.get('LEDGER_UPLOAD_TOKEN') or '').strip()
    url = _sanitize_url(body.get('url') or form.get('url') or request.args.get('url') or 'https://spzs.scjgj.sh.gov.cn/p4/api/v1/data/upload/file')
    cookie = (body.get('cookie') or form.get('cookie') or request.args.get('cookie') or '').strip()
    user_agent = (body.get('user_agent') or form.get('user_agent') or request.args.get('user_agent') or 'Apifox/1.0.0 (https://apifox.com)').strip()
    # 业务必填字段，日期默认今天，其它按要求维持不变
    from datetime import datetime
    data_date = (body.get('dataDate') or form.get('dataDate') or request.args.get('dataDate') or datetime.today().strftime('%Y-%m-%d'))
    field_name = (body.get('fieldName') or form.get('fieldName') or request.args.get('fieldName') or 'catering')
    license_no = (body.get('license') or form.get('license') or request.args.get('license') or 'JY23101140040138')
    biz_loc_id = (body.get('businessLocationId') or form.get('businessLocationId') or request.args.get('businessLocationId') or '11215')
    if not path:
      return jsonify({'error': '缺少 path'}), 400
    if not os.path.isfile(path):
      return jsonify({'error': '文件不存在', 'path': path}), 404
    if not token:
      return jsonify({'error': '缺少 token'}), 400
    import requests
    from datetime import datetime as _dt
    fn = os.path.basename(path)
    try:
      fn = f"上海萨莉亚餐饮有限公司_餐饮_{_dt.today().strftime('%Y%m%d')}.xlsx"
    except Exception:
      pass
    ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    with open(path, 'rb') as f:
      files = [('file', (fn, f, ct))]
      headers = {
        'Authorization': f'Bearer {token}',
        'User-Agent': user_agent,
        'Accept': '*/*',
      }
      if cookie:
        headers['Cookie'] = cookie
      payload = {
        'dataDate': str(data_date),
        'fieldName': str(field_name),
        'license': str(license_no),
        'businessLocationId': str(biz_loc_id),
      }
      r = requests.post(url, data=payload, files=files, headers=headers, timeout=120)
    try:
      data = r.json()
    except Exception:
      data = {'raw': r.text}
    ok = (r.status_code == 200) and bool((data or {}).get('success'))
    return jsonify({'statusCode': r.status_code, 'ok': ok, 'upstream': data})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.get('/api/v1/ledger/external-upload/status')
def ledger_external_upload_status():
  try:
    import requests
    from datetime import datetime
    token = (request.args.get('token') or os.environ.get('LEDGER_UPLOAD_TOKEN') or '').strip()
    if not token:
      return jsonify({'error': '缺少 token'}), 400
    full = _sanitize_url(request.args.get('full_url'))
    url = _sanitize_url(request.args.get('url') or 'https://spzs.scjgj.sh.gov.cn/p4/api/v1/data/upload/list')
    cookie = (request.args.get('cookie') or '').strip()
    user_agent = (request.args.get('user_agent') or 'Apifox/1.0.0 (https://apifox.com)').strip()
    host_hdr = (request.args.get('host') or 'spzs.scjgj.sh.gov.cn').strip()
    params = {
      'dataDate': (request.args.get('dataDate') or datetime.today().strftime('%Y-%m-%d')),
      'fieldNameDisplay': request.args.get('fieldNameDisplay') or '餐饮',
      'enterpriseName': request.args.get('enterpriseName') or '上海萨莉亚餐饮有限公司墨玉南路店',
      'license': request.args.get('license') or 'JY23101140040138',
      'fieldName': request.args.get('fieldName') or 'catering',
      'businessLocationId': request.args.get('businessLocationId') or '11215',
      'childEnterpriseId': request.args.get('childEnterpriseId') or '',
      'loading': request.args.get('loading') or 'true'
    }
    headers = {
      'Authorization': f'Bearer {token}',
      'User-Agent': user_agent,
      'Accept': '*/*',
    }
    if cookie:
      headers['Cookie'] = cookie
    if full:
      r = requests.get(full, headers=headers, timeout=120)
    else:
      r = requests.get(url, params=params, headers=headers, timeout=120)
    try:
      data = r.json()
    except Exception:
      data = {'raw': r.text}
    # 计算整体成功
    status = None
    failed = None
    try:
      arr = (data or {}).get('content') or []
      status = (arr[0] or {}).get('status') if arr else None
      sumobj = (data or {}).get('sum') or {}
      failed = sumobj.get('failed')
    except Exception:
      pass
    ok = (status == 'success') and (failed == 0)
    return jsonify({'statusCode': r.status_code, 'ok': ok, 'status': status, 'sum': (data or {}).get('sum'), 'upstream': data})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

def _fill_delivery_formulas(wb):
  try:
    target = None
    for cand in wb.worksheets:
      title = str(getattr(cand, 'title', '') or '')
      if title == '配送表' or ('配送' in title):
        target = cand
        break
    if target is None:
      return
    sh = target
    hdr_row = None
    for r in range(1, min(15, sh.max_row+1)):
      names = [str(sh.cell(row=r, column=c).value or '').strip() for c in range(1, min(40, sh.max_column+1))]
      if any(('产品代码' in x or '统一社会信用' in x or '生产厂商' in x) for x in names):
        hdr_row = r
        break
    if hdr_row is None:
      hdr_row = 1
    headers = {}
    for c in range(1, sh.max_column+1):
      name = str(sh.cell(row=hdr_row, column=c).value or '').strip()
      if name:
        headers[name] = c
    def _col(*keys):
      for k in keys:
        for name, col in headers.items():
          if k in name:
            return col
      return None
    col_prod_code = _col('产品代码', '商品编码', '编码', '产品编号')
    col_uscc = _col('统一社会信用代码', '统一社会信用码', '社会信用代码')
    col_manu = _col('生产厂商', '生产厂家', '生产企业', '厂家')
    start = hdr_row + 1
    end = min(sh.max_row, hdr_row + 499)
    for r in range(start, end+1):
      try:
        if col_prod_code:
          sh.cell(row=r, column=col_prod_code).value = f'=VLOOKUP($A{r},产品信息!A:O,2,FALSE)&""'
        if col_uscc:
          sh.cell(row=r, column=col_uscc).value = f'=VLOOKUP($F{r},购货者!A:D,2,FALSE)&""'
        if col_manu:
          sh.cell(row=r, column=col_manu).value = f'=VLOOKUP($A{r},产品信息!A:P,11,FALSE)&""'
      except Exception:
        pass
  except Exception:
    pass

@app.post('/api/v1/ledger/process')
def ledger_process():
  try:
    body = request.json or {}
    form = request.form or {}
    ct = str(request.headers.get('Content-Type') or '').lower()
    if ('multipart/form-data' in ct) or (request.files):
      return ledger_process_upload()
    at, ac = _ledger_active_paths()
    template_path = body.get('template_path') or at or os.path.join(os.path.dirname(__file__), '..', '上海萨莉亚餐饮有限公司_餐饮_202511.xlsx')
    conversion_path = body.get('conversion_path') or ac or os.path.join(os.path.dirname(__file__), '..', '食品台账换算千克数 202510.xlsx')
    dry_run = bool(body.get('dry_run')) or bool((form.get('dry_run') or '').strip())
    items = body.get('items') or []
    print(f"items:{items}")
    if not items and ('items' in form):
      import json as _json
      try:
        items_txt = (form.get('items') or '').strip()
        items = _json.loads(items_txt) if items_txt else []
        if not isinstance(items, list):
          items = []
      except Exception:
        items = []
    if load_workbook is None:
      return jsonify({'error': '需要openpyxl依赖'}), 500
    conv_map = _read_conversion_table(conversion_path)
    def _norm_code(s):
      if s is None:
        return ''
      s = str(s).strip()
      d = ''.join(ch for ch in s if ch.isdigit())
      if d:
        d = d.lstrip('0')
        return d or '0'
      return s
    def _is_code_like_raw(s):
      t = str(s or '').strip()
      if not t:
        return False
      digits = ''.join(ch for ch in t if ch.isdigit())
      others = ''.join(ch for ch in t if (not ch.isdigit()) and ch not in [' ', '\t', '.', '-', '/'])
      if others:
        return False
      return 5 <= len(digits) <= 8
    conv_map_norm = { _norm_code(k): v for k, v in conv_map.items() }
    wb = load_workbook(template_path, data_only=True)
    rec_map = {}
    rec_map_norm = {}
    for it in items:
      code = str((it or {}).get('product_code') or '').strip()
      if not code:
        continue
      rec_map[code] = {
        'box': it.get('box_count'),
        'piece': it.get('piece_count'),
        'prod': it.get('production_date')
      }
      rec_map_norm[_norm_code(code)] = rec_map[code]
    target = None
    def _auto_code_col_with(cand_sh, hdr_row_in):
      best_col = None
      best_hits = 0.0
      for c in range(1, cand_sh.max_column+1):
        hits = 0.0
        for r in range(hdr_row_in+1, cand_sh.max_row+1):
          val = str(cand_sh.cell(row=r, column=c).value or '').strip()
          if not val:
            continue
          if not _is_code_like_raw(val):
            continue
          vn = _norm_code(val)
          if vn in rec_map_norm or vn in conv_map_norm:
            hits += 1.0
        if hits > best_hits:
          best_hits = hits
          best_col = c
      return best_col if best_hits >= 2.0 else None

    preferred = None
    for cand in wb.worksheets:
      title = str(getattr(cand, 'title', '') or '')
      if title == '进货表' or ('进货' in title):
        preferred = cand
        break
    if preferred is not None:
      hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg = _locate_headers(preferred)
      if col_code is None:
        col_code = _auto_code_col_with(preferred, hdr_row)
      if col_code is not None:
        target = (preferred, hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg)
    if target is None:
      for cand in wb.worksheets:
        hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg = _locate_headers(cand)
        if col_code is None:
          col_code = _auto_code_col_with(cand, hdr_row)
        if col_code is not None:
          target = (cand, hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg)
          break
    if target is None:
      return jsonify({'error': '模板缺少商品编码列'}), 500
    sh, hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg = target
    if col_qty is None or col_in_date is None or col_prod_date is None:
      missing = []
      if col_qty is None:
        missing.append('采购数量')
      if col_in_date is None:
        missing.append('进货日期')
      if col_prod_date is None:
        missing.append('生产日期')
      return {'error': f"模板缺少必需列: {','.join(missing)}"}, 500
    now = datetime.now()
    today_date = now.date()
    yest_date = (now - timedelta(days=1)).date()
    errors = []
    del_rows = []
    updated_rows = 0
    written_rows = []
    unmatched_codes = []
    conv_missing_codes = []
    audit_trail = []
    validate_pass = 0
    validate_fail = 0
    validate_mismatches = []
    for r in range(hdr_row+1, sh.max_row+1):
      code_raw = str(sh.cell(row=r, column=col_code).value or '').strip()
      if not code_raw:
        del_rows.append(r)
        continue
      code_norm = _norm_code(code_raw)
      if not _is_code_like_raw(code_raw):
        unmatched_codes.append({'row': r, 'codeRaw': code_raw, 'codeNorm': code_norm})
        errors.append({'code': code_raw, 'error': '单元格内容非编码格式'})
        continue
      rec = rec_map.get(code_raw) or rec_map_norm.get(code_norm)
      raw_qty = None
      if code_raw.startswith('72') or code_norm.startswith('72'):
        raw_qty = 1
      else:
        if rec is None:
          errors.append({'code': code_raw, 'error': '未找到对应记录'})
          unmatched_codes.append({'row': r, 'codeRaw': code_raw, 'codeNorm': code_norm})
          continue
        piece_val = rec.get('piece')
        box_val = rec.get('box')
        if piece_val is not None:
          try:
            q = float(str(piece_val))
            raw_qty = q if q > 0 else None
          except Exception:
            raw_qty = None
        if raw_qty is None and box_val is not None:
          try:
            q = float(str(box_val))
            raw_qty = q if q > 0 else None
          except Exception:
            raw_qty = None
      if raw_qty is None:
        errors.append({'code': code_raw, 'error': '数量为空或为0'})
        continue
      result_qty = raw_qty
      sh.cell(row=r, column=col_qty).value = result_qty
      try:
        sh.cell(row=r, column=col_qty).number_format = '0.##'
      except Exception:
        pass
      sh.cell(row=r, column=col_in_date).value = today_date
      try:
        sh.cell(row=r, column=col_in_date).number_format = 'yyyy-mm-dd'
      except Exception:
        pass
      pv = None
      if code_raw.startswith('72') or code_norm.startswith('72'):
        pv = yest_date
      else:
        if rec and rec.get('prod'):
          try:
            from datetime import datetime as _dt
            s = str(rec.get('prod'))
            s = s.replace('.', '-').replace('/', '-')
            pv = _dt.strptime(s[:10], '%Y-%m-%d').date()
          except Exception:
            pv = None
      sh.cell(row=r, column=col_prod_date).value = pv or ''
      try:
        if pv:
          sh.cell(row=r, column=col_prod_date).number_format = 'yyyy-mm-dd'
      except Exception:
        pass
      audit_trail.append({'row': r, 'code': code_raw, 'rawQty': raw_qty, 'writtenQty': result_qty})
      updated_rows += 1
      written_rows.append(r)
    strict_keep = _as_bool(body.get('strict_keep_written_only'), None)
    strict_keep = _as_bool(form.get('strict_keep_written_only'), strict_keep)
    if strict_keep is None:
      strict_keep = True
    if strict_keep:
      all_rows = list(range(hdr_row+1, sh.max_row+1))
      del_rows = [idx for idx in all_rows if idx not in written_rows]
      for idx in sorted(del_rows, reverse=True):
        sh.delete_rows(idx, 1)
    else:
      del_rows = []
    # 写入配送表关键列公式
    try:
      if not dry_run:
        _fill_delivery_formulas(wb)
    except Exception:
      pass
    # allow explicit output path to avoid permission issues
    def _default_out_path(tpl_path):
      base_dir = os.path.dirname(tpl_path)
      gen_dir = os.path.join(UPLOAD_DIR, 'generated')
      _ensure_dir(gen_dir)
      ymd = now.strftime('%Y%m%d')
      fname = f"{ymd}-上海萨莉亚餐饮有限公司_餐饮_{ymd}.xlsx"
      return os.path.join(gen_dir, fname)
    out_path = (body.get('output_path') or form.get('output_path') or '').strip()
    if not out_path:
      out_path = _default_out_path(template_path)
    try:
      if out_path:
        _ensure_dir(os.path.dirname(out_path))
        wb.save(out_path)
        return jsonify({'saved': True, 'path': os.path.abspath(out_path), 'errors': errors, 'unmatchedCodes': unmatched_codes, 'convMissingCodes': conv_missing_codes, 'sheetName': str(getattr(sh, 'title', '') or ''), 'usedOutputPath': True, 'auditTrail': audit_trail, 'validation': {'pass': validate_pass, 'fail': validate_fail, 'mismatches': validate_mismatches}, 'meta': {'updatedRows': updated_rows, 'deletedRows': len(del_rows), 'totalRows': updated_rows + len(del_rows), 'itemsCount': len(items), 'wasDryRun': False}})
      # default save to generated file rather than overwrite template
      wb.save(out_path)
      return jsonify({'saved': True, 'path': os.path.abspath(out_path), 'errors': errors, 'unmatchedCodes': unmatched_codes, 'convMissingCodes': conv_missing_codes, 'sheetName': str(getattr(sh, 'title', '') or ''), 'usedOutputPath': True, 'auditTrail': audit_trail, 'validation': {'pass': validate_pass, 'fail': validate_fail, 'mismatches': validate_mismatches}, 'meta': {'updatedRows': updated_rows, 'deletedRows': len(del_rows), 'totalRows': updated_rows + len(del_rows), 'itemsCount': len(items), 'wasDryRun': False}})
    except Exception as e:
      try:
        dirn = os.path.dirname(template_path)
        base = os.path.splitext(os.path.basename(template_path))[0]
        alt = os.path.join(dirn, base + '_out.xlsx')
        wb.save(alt)
        return jsonify({'saved': True, 'path': os.path.abspath(alt), 'errors': errors, 'fallback': True, 'errorMsg': str(e), 'updatedRows': updated_rows, 'deletedRows': len(del_rows), 'unmatchedCodes': unmatched_codes, 'convMissingCodes': conv_missing_codes, 'sheetName': str(getattr(sh, 'title', '') or ''), 'auditTrail': audit_trail, 'validation': {'pass': validate_pass, 'fail': validate_fail, 'mismatches': validate_mismatches}})
      except Exception as e2:
        try:
          gen_dir = os.path.join(UPLOAD_DIR, 'generated')
          _ensure_dir(gen_dir)
          safe_name = (os.path.splitext(os.path.basename(template_path))[0] or 'ledger') + '_out.xlsx'
          alt2 = os.path.join(gen_dir, safe_name)
          wb.save(alt2)
          return jsonify({'saved': True, 'path': os.path.abspath(alt2), 'errors': errors, 'fallback': True, 'errorMsg': f"{e}; {e2}", 'updatedRows': updated_rows, 'deletedRows': len(del_rows), 'unmatchedCodes': unmatched_codes, 'convMissingCodes': conv_missing_codes, 'sheetName': str(getattr(sh, 'title', '') or ''), 'auditTrail': audit_trail, 'validation': {'pass': validate_pass, 'fail': validate_fail, 'mismatches': validate_mismatches}})
        except Exception as e3:
          return jsonify({'error': f'save failed: {e}; fallback failed: {e2}; generated failed: {e3}'}), 500
  except Exception as e:
    return jsonify({'error': str(e)}), 500

def _process_ledger_with_paths(template_path, conversion_path, items, output_path, dry_run=False, strict_keep_written_only=True):
  try:
    if load_workbook is None:
      return {'error': '需要openpyxl依赖'}, 500
    conv_map = _read_conversion_table(conversion_path)
    def _norm_code(s):
      if s is None:
        return ''
      s = str(s).strip()
      d = ''.join(ch for ch in s if ch.isdigit())
      if d:
        d = d.lstrip('0')
        return d or '0'
      return s
    def _is_code_like_raw(s):
      t = str(s or '').strip()
      if not t:
        return False
      digits = ''.join(ch for ch in t if ch.isdigit())
      others = ''.join(ch for ch in t if (not ch.isdigit()) and ch not in [' ', '\t', '.', '-', '/'])
      if others:
        return False
      return 5 <= len(digits) <= 8
    conv_map_norm = { _norm_code(k): v for k, v in conv_map.items() }
    wb = load_workbook(template_path, data_only=True)
    rec_map = {}
    rec_map_norm = {}
    for it in items or []:
      code = str((it or {}).get('product_code') or '').strip()
      if not code:
        continue
      rec_map[code] = {
        'box': it.get('box_count'),
        'piece': it.get('piece_count'),
        'prod': it.get('production_date')
      }
      rec_map_norm[_norm_code(code)] = rec_map[code]
    target = None
    def _auto_code_col_with(cand_sh, hdr_row_in):
      best_col = None
      best_hits = 0.0
      for c in range(1, cand_sh.max_column+1):
        hits = 0.0
        for r in range(hdr_row_in+1, cand_sh.max_row+1):
          val = str(cand_sh.cell(row=r, column=c).value or '').strip()
          if not val:
            continue
          if not _is_code_like_raw(val):
            continue
          vn = _norm_code(val)
          if vn in rec_map_norm or vn in conv_map_norm:
            hits += 1.0
        if hits > best_hits:
          best_hits = hits
          best_col = c
      return best_col if best_hits >= 2.0 else None
    preferred = None
    for cand in wb.worksheets:
      title = str(getattr(cand, 'title', '') or '')
      if title == '进货表' or ('进货' in title):
        preferred = cand
        break
    if preferred is not None:
      hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg = _locate_headers(preferred)
      if col_code is None:
        col_code = _auto_code_col_with(preferred, hdr_row)
      if col_code is not None:
        target = (preferred, hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg)
    if target is None:
      for cand in wb.worksheets:
        hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg = _locate_headers(cand)
        if col_code is None:
          col_code = _auto_code_col_with(cand, hdr_row)
        if col_code is not None:
          target = (cand, hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg)
          break
    if target is None:
      return {'error': '模板缺少商品编码列'}, 500
    sh, hdr_row, col_code, col_qty, col_in_date, col_prod_date, col_kg = target
    if col_qty is None or col_in_date is None or col_prod_date is None:
      missing = []
      if col_qty is None:
        missing.append('采购数量')
      if col_in_date is None:
        missing.append('进货日期')
      if col_prod_date is None:
        missing.append('生产日期')
      return {'error': f"模板缺少必需列: {','.join(missing)}"}, 500
    now = datetime.now()
    today_date = now.date()
    yest_date = (now - timedelta(days=1)).date()
    errors = []
    del_rows = []
    updated_rows = 0
    written_rows = []
    unmatched_codes = []
    conv_missing_codes = []
    audit_trail = []
    validate_pass = 0
    validate_fail = 0
    validate_mismatches = []
    for r in range(hdr_row+1, sh.max_row+1):
      code_raw = str(sh.cell(row=r, column=col_code).value or '').strip()
      if not code_raw:
        del_rows.append(r)
        continue
      code_norm = _norm_code(code_raw)
      if not _is_code_like_raw(code_raw):
        unmatched_codes.append({'row': r, 'codeRaw': code_raw, 'codeNorm': code_norm})
        errors.append({'code': code_raw, 'error': '单元格内容非编码格式'})
        continue
      rec = rec_map.get(code_raw) or rec_map_norm.get(code_norm)
      raw_qty = None
      if code_raw.startswith('72') or code_norm.startswith('72'):
        raw_qty = 1
      else:
        if rec is None:
          errors.append({'code': code_raw, 'error': '未找到对应记录'})
          unmatched_codes.append({'row': r, 'codeRaw': code_raw, 'codeNorm': code_norm})
          continue
        piece_val = rec.get('piece')
        box_val = rec.get('box')
        if piece_val is not None:
          try:
            q = float(str(piece_val))
            raw_qty = q if q > 0 else None
          except Exception:
            raw_qty = None
        if raw_qty is None and box_val is not None:
          try:
            q = float(str(box_val))
            raw_qty = q if q > 0 else None
          except Exception:
            raw_qty = None
      if raw_qty is None:
        errors.append({'code': code_raw, 'error': '数量为空或为0'})
        continue
      result_qty = raw_qty
      sh.cell(row=r, column=col_qty).value = result_qty
      try:
        sh.cell(row=r, column=col_qty).number_format = '0.##'
      except Exception:
        pass
      sh.cell(row=r, column=col_in_date).value = today_date
      try:
        sh.cell(row=r, column=col_in_date).number_format = 'yyyy-mm-dd'
      except Exception:
        pass
      pv = None
      if code_raw.startswith('72') or code_norm.startswith('72'):
        pv = yest_date
      else:
        if rec and rec.get('prod'):
          try:
            from datetime import datetime as _dt
            s = str(rec.get('prod'))
            s = s.replace('.', '-').replace('/', '-')
            pv = _dt.strptime(s[:10], '%Y-%m-%d').date()
          except Exception:
            pv = None
      sh.cell(row=r, column=col_prod_date).value = pv or ''
      try:
        if pv:
          sh.cell(row=r, column=col_prod_date).number_format = 'yyyy-mm-dd'
      except Exception:
        pass
      audit_trail.append({'row': r, 'code': code_raw, 'rawQty': raw_qty, 'writtenQty': result_qty})
      updated_rows += 1
      written_rows.append(r)
    if strict_keep_written_only:
      all_rows = list(range(hdr_row+1, sh.max_row+1))
      del_rows = [idx for idx in all_rows if idx not in written_rows]
      for idx in sorted(del_rows, reverse=True):
        sh.delete_rows(idx, 1)
    else:
      del_rows = []
    try:
      if not dry_run:
        _fill_delivery_formulas(wb)
    except Exception:
      pass
    total_rows = updated_rows + len(del_rows)
    meta = {'updatedRows': updated_rows, 'deletedRows': len(del_rows), 'totalRows': total_rows}
    if dry_run:
      return {'saved': False, 'preview': True, 'errors': errors, 'unmatchedCodes': unmatched_codes, 'convMissingCodes': conv_missing_codes, 'sheetName': str(getattr(sh, 'title', '') or ''), 'auditTrail': audit_trail, 'validation': {'pass': validate_pass, 'fail': validate_fail, 'mismatches': validate_mismatches}, 'meta': meta}, 200
    out_path = (output_path or '').strip()
    try:
      if out_path:
        _ensure_dir(os.path.dirname(out_path))
        wb.save(out_path)
        return {'saved': True, 'path': os.path.abspath(out_path), 'errors': errors, 'unmatchedCodes': unmatched_codes, 'convMissingCodes': conv_missing_codes, 'sheetName': str(getattr(sh, 'title', '') or ''), 'usedOutputPath': True, 'auditTrail': audit_trail, 'validation': {'pass': validate_pass, 'fail': validate_fail, 'mismatches': validate_mismatches}, 'meta': meta}, 200
      wb.save(template_path)
      return {'saved': True, 'path': os.path.abspath(template_path), 'errors': errors, 'unmatchedCodes': unmatched_codes, 'convMissingCodes': conv_missing_codes, 'sheetName': str(getattr(sh, 'title', '') or ''), 'usedOutputPath': False, 'auditTrail': audit_trail, 'validation': {'pass': validate_pass, 'fail': validate_fail, 'mismatches': validate_mismatches}, 'meta': meta}, 200
    except Exception as e:
      try:
        dirn = os.path.dirname(template_path)
        base = os.path.splitext(os.path.basename(template_path))[0]
        alt = os.path.join(dirn, base + '_out.xlsx')
        wb.save(alt)
        return {'saved': True, 'path': os.path.abspath(alt), 'errors': errors, 'fallback': True, 'errorMsg': str(e), 'unmatchedCodes': unmatched_codes, 'convMissingCodes': conv_missing_codes, 'sheetName': str(getattr(sh, 'title', '') or ''), 'auditTrail': audit_trail, 'validation': {'pass': validate_pass, 'fail': validate_fail, 'mismatches': validate_mismatches}, 'meta': meta}, 200
      except Exception as e2:
        try:
          gen_dir = os.path.join(UPLOAD_DIR, 'generated')
          _ensure_dir(gen_dir)
          safe_name = (os.path.splitext(os.path.basename(template_path))[0] or 'ledger') + '_out.xlsx'
          alt2 = os.path.join(gen_dir, safe_name)
          wb.save(alt2)
          return {'saved': True, 'path': os.path.abspath(alt2), 'errors': errors, 'fallback': True, 'errorMsg': f"{e}; {e2}", 'unmatchedCodes': unmatched_codes, 'convMissingCodes': conv_missing_codes, 'sheetName': str(getattr(sh, 'title', '') or ''), 'auditTrail': audit_trail, 'validation': {'pass': validate_pass, 'fail': validate_fail, 'mismatches': validate_mismatches}, 'meta': meta}, 200
        except Exception as e3:
          return {'error': f'save failed: {e}; fallback failed: {e2}; generated failed: {e3}'}, 500

  except Exception as e:
    return {'error': str(e)}, 500

def _read_conversion_table_csv(csv_path):
  import csv
  m = {}
  try:
    with open(csv_path, 'r', encoding='utf-8') as f:
      reader = csv.reader(f)
      rows = list(reader)
    hdr = rows[0] if rows else []
    def _find_idx(cols):
      for i, name in enumerate(hdr):
        s = str(name or '')
        for k in cols:
          if k in s:
            return i
      return None
    idx_code = _find_idx(['商品编码','编码','商品编号','代码','货品编码','产品代码'])
    idx_unitkg = _find_idx(['进货单位千克','单位千克','kg','千克'])
    if idx_code is None or idx_unitkg is None:
      return {}
    for r in rows[1:]:
      code = str(r[idx_code] or '').strip()
      if not code:
        continue
      try:
        uk = float(str(r[idx_unitkg]).strip())
      except Exception:
        uk = 0.0
      m[code] = uk
  except Exception:
    return {}
  return m

@app.post('/api/v1/ledger/process-upload')
def ledger_process_upload():
  try:
    if 'template' not in request.files:
      # allow using active template if not uploaded
      at, ac = _ledger_active_paths()
      if not at:
        return jsonify({'error': '缺少进货表文件'}), 400
      template_file = None
      conversion_file = request.files.get('conversion')
      items_json = (request.form.get('items') or '').strip()
      dry_run = _as_bool(request.form.get('dry_run'), False)
      output_path = (request.form.get('output_path') or '').strip()
      import json
      try:
        items = json.loads(items_json) if items_json else []
        if not isinstance(items, list):
          items = []
      except Exception:
        items = []
      conv_path = None
      if conversion_file:
        _ensure_dir(UPLOAD_DIR)
        cname = conversion_file.filename or 'conversion.xlsx'
        cpath = os.path.join(UPLOAD_DIR, cname)
        conversion_file.save(cpath)
        conv_path = cpath
      if conv_path is None:
        conv_path = ac or os.path.join(os.path.dirname(__file__), '..', '食品台账换算千克数 202510.xlsx')
      if not output_path:
        from datetime import datetime
        now = datetime.now()
        ymd = now.strftime('%Y%m%d')
        gen_dir = os.path.join(UPLOAD_DIR, 'generated')
        _ensure_dir(gen_dir)
        output_path = os.path.join(gen_dir, f"上海萨莉亚餐饮有限公司_餐饮_{ymd}.xlsx")
      resp, code = _process_ledger_with_paths(at, conv_path, items, output_path, dry_run, _as_bool(request.form.get('strict_keep_written_only'), True))
      return jsonify(resp), code
    template_file = request.files['template']
    conversion_file = request.files.get('conversion')
    items_json = (request.form.get('items') or '').strip()
    output_path = (request.form.get('output_path') or '').strip()
    import json
    items = []
    try:
      items = json.loads(items_json) if items_json else []
      if not isinstance(items, list):
        items = []
    except Exception:
      items = []
    _ensure_dir(UPLOAD_DIR)
    tname = template_file.filename or 'template.xlsx'
    ext_ok = any(tname.lower().endswith(x) for x in ['.xlsx','.xls'])
    if not ext_ok:
      return jsonify({'error': '进货表文件必须为Excel格式'}), 400
    tpath = os.path.join(UPLOAD_DIR, tname)
    template_file.save(tpath)
    conv_path = None
    conv_map_extra = {}
    if conversion_file:
      cname = conversion_file.filename or 'conversion.xlsx'
      cpath = os.path.join(UPLOAD_DIR, cname)
      conversion_file.save(cpath)
      conv_path = cpath
      if cname.lower().endswith('.csv'):
        conv_map_extra = _read_conversion_table_csv(cpath)
    if conv_path is None:
      at, ac = _ledger_active_paths()
      conv_path = ac or os.path.join(os.path.dirname(__file__), '..', '食品台账换算千克数 202510.xlsx')
    # persist uploaded files and activate
    st = _read_ledger_store()
    files = st.get('files') or {}
    tpl_list = files.get('templates') or []
    conv_list = files.get('conversions') or []
    from datetime import datetime
    tpl_entry = {'name': tname, 'path': tpath, 'uploadedAt': datetime.utcnow().isoformat()}
    tpl_list = [e for e in tpl_list if e.get('path') != tpath] + [tpl_entry]
    st['active_template'] = tpath
    files['templates'] = tpl_list
    if conv_path:
      conv_entry = {'name': os.path.basename(conv_path), 'path': conv_path, 'uploadedAt': datetime.utcnow().isoformat()}
      conv_list = [e for e in conv_list if e.get('path') != conv_path] + [conv_entry]
      files['conversions'] = conv_list
      st['active_conversion'] = conv_path
    st['files'] = files
    _write_ledger_store(st)
    dry_run = _as_bool(request.form.get('dry_run'), False)
    if not output_path:
      from datetime import datetime
      now = datetime.now()
      ymd = now.strftime('%Y%m%d')
      gen_dir = os.path.join(UPLOAD_DIR, 'generated')
      _ensure_dir(gen_dir)
      output_path = os.path.join(gen_dir, f"上海萨莉亚餐饮有限公司_餐饮_{ymd}.xlsx")
    resp, code = _process_ledger_with_paths(tpath, conv_path, items, output_path, dry_run, _as_bool(request.form.get('strict_keep_written_only'), True))
    return jsonify(resp), code
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.get('/api/v1/ledger/files')
def ledger_files_list():
  try:
    st = _read_ledger_store()
    return jsonify({'active_template': st.get('active_template'), 'active_conversion': st.get('active_conversion'), 'files': st.get('files') or {}})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.post('/api/v1/ledger/files/activate')
def ledger_files_activate():
  try:
    body = request.json or {}
    kind = (body.get('kind') or '').strip()
    path = (body.get('path') or '').strip()
    if kind not in ['template','conversion'] or not path:
      return jsonify({'error': '参数错误'}), 400
    if not os.path.isfile(path):
      return jsonify({'error': '文件不存在'}), 404
    st = _read_ledger_store()
    files = st.get('files') or {}
    arr = files.get('templates' if kind=='template' else 'conversions') or []
    if not any((e.get('path') == path) for e in arr):
      from datetime import datetime
      ent = {'name': os.path.basename(path), 'path': path, 'uploadedAt': datetime.utcnow().isoformat()}
      arr.append(ent)
      files['templates' if kind=='template' else 'conversions'] = arr
    st['files'] = files
    if kind == 'template':
      st['active_template'] = path
    else:
      st['active_conversion'] = path
    _write_ledger_store(st)
    return jsonify({'ok': True})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

_rate = {}
_lock = threading.Lock()

def _rate_limit(key, max_per_min=20):
  now = time.time()
  with _lock:
    arr = _rate.get(key) or []
    arr = [t for t in arr if now - t < 60]
    if len(arr) >= max_per_min:
      return False
    arr.append(now)
    _rate[key] = arr
    return True

@app.post('/api/v1/ledger/image-recognize')
def ledger_image_recognize():
  try:
    ip = request.remote_addr or 'unknown'
    if not _rate_limit(f'img:{ip}', 30):
      return jsonify({'error': 'rate limited'}), 429
    if 'file' not in request.files:
      return jsonify({'error': '缺少图片文件'}), 400
    f = request.files['file']
    b = f.read()
    if not b:
      return jsonify({'error': '空文件'}), 400
    form = request.form or {}
    key = 'sk-N4pVo05nxeab2N09OOfPV3SipT48319L7kt7vfOSGXRcP2KT' or form.get('api_key') or os.environ.get('OPENAI_API_KEY')
    api_url = 'https://api.muhuo.site/v1/chat/completions' or form.get('api_url') or os.environ.get('OPENAI_API_URL')
    api_url = str(api_url or '').replace('`','').strip()
    if api_url.endswith('/chat/completion'):
      api_url = api_url[:-1] + 's'
    items = []
    dbg = {'api_url': api_url, 'key_present': bool(key)}
    print(f"key:{key}, apu_url:{api_url}")
    if key:
      import requests
      mime = 'image/jpeg'
      try:
        from PIL import Image
        import io
        im = Image.open(io.BytesIO(b)).convert('RGB')
        w, h = im.size
        max_w = 1200
        if w > max_w:
          ratio = max_w / float(w)
          im = im.resize((int(w * ratio), int(h * ratio)))
        buf = io.BytesIO()
        im.save(buf, format='JPEG', quality=75)
        b = buf.getvalue()
      except Exception:
        sig = b[:8]
        if sig.startswith(b'\x89PNG'):
          mime = 'image/png'
          ext = '.png'
        elif sig.startswith(b'\xff\xd8'):
          mime = 'image/jpeg'
          ext = '.jpg'
        else:
          # 无法识别类型，回退为 data URL（不保存为 .bin）
          img_b64 = base64.b64encode(b).decode('ascii')
          image_http_url = f"data:{mime};base64,{img_b64}"
          save_ok = False
          ext = None
      import time, random
      if mime == 'image/png':
        fname = f"img_{int(time.time())}_{random.randint(1000,9999)}.png"
      elif mime == 'image/jpeg':
        fname = f"img_{int(time.time())}_{random.randint(1000,9999)}.jpg"
      else:
        # 已经回退为 data URL，不再保存不可识别的 .bin 文件
        fname = f"img_{int(time.time())}_{random.randint(1000,9999)}.jpg"
      fpath = os.path.join(IMAGES_DIR, fname)
      save_ok = True
      try:
        with open(fpath, 'wb') as out:
          out.write(b)
      except Exception as save_e:
        save_ok = False
        print({'AI_IMAGE_SAVE_ERROR': {'path': fpath, 'error': str(save_e)}})
      if save_ok:
        base_url_env = os.environ.get('PUBLIC_BASE_URL')
        if base_url_env:
          base_url = base_url_env.strip().rstrip('/')
        else:
          xf_proto = (request.headers.get('X-Forwarded-Proto') or '').strip()
          xf_host = (request.headers.get('X-Forwarded-Host') or '').strip()
          if xf_host:
            scheme = xf_proto if xf_proto in ['http','https'] else 'https'
            base_url = f"{scheme}://{xf_host}"
          else:
            base_url = (request.url_root or request.host_url or '').strip().rstrip('/')
        base_url = str(base_url or '').replace('`','').strip()
        image_http_url = f"{base_url}/uploads/images/{fname}"
        image_http_url = str(image_http_url or '').replace('`','').strip()
        print({'AI_IMAGE_URL': image_http_url})
      else:
        img_b64 = base64.b64encode(b).decode('ascii')
        image_http_url = f"data:{mime};base64,{img_b64}"
      prompt = '识别送货单中的商品编码、拆零数、箱数和生产日期，返回JSON数组，键为product_code,piece_count,box_count,production_date，日期YYYY-MM-DD。'
      payload = {
        'model': 'gemini-2.5-pro',
        'messages': [
          {'role': 'system', 'content': '你是一个严谨的表格信息抽取助手，只返回结构化JSON，不要多余文本。'},
          {'role': 'user', 'content': [
            {'type': 'text', 'text': prompt},
            {'type': 'image_url', 'image_url': {'url': image_http_url, 'detail': 'low'}}
          ]}
        ],
        'temperature': 0,
        'stream': True
      }
      verbose = str(form.get('verbose') or request.args.get('verbose') or os.environ.get('LEDGER_DEBUG') or '').strip()
      debug = {'api_url': api_url, 'model': payload.get('model'), 'image_size': len(b), 'stream': True}
      try:
        r = requests.post(api_url, json=payload, headers={'Authorization': f'Bearer {key}', 'Content-Type': 'application/json'}, stream=True, timeout=1200)
      except Exception as req_e:
        debug['request_error'] = str(req_e)
        print({'AI_REQUEST_ERROR': debug})
        return jsonify({'error': 'upstream_request_failed', 'debug': debug}), 502
      debug['status_code'] = r.status_code
      if verbose:
        print({'AI_UPSTREAM_OK': {'status_code': r.status_code}})
      if r.status_code == 200:
        import json as _json
        text_parts = []
        raw_parts = []
        for line in r.iter_lines(decode_unicode=True):
          if not line:
            continue
          s = str(line)
          if s.startswith('data:'):
            s = s[5:].strip()
          if s == '[DONE]':
            break
          try:
            evt = _json.loads(s)
            chs = evt.get('choices') or []
            ct = ''
            if chs:
              d = chs[0] or {}
              ct = (d.get('delta') or {}).get('content') or (d.get('message') or {}).get('content') or ''
            else:
              ct = evt.get('content') or ''
            if isinstance(ct, list):
              ct = ''.join([ (p.get('text') if isinstance(p, dict) else str(p)) or '' for p in ct ])
            if ct:
              text_parts.append(str(ct))
          except Exception:
            raw_parts.append(s)
        text = ''.join(text_parts) or ''.join(raw_parts)
        try:
          import json, re
          s = str(text or '')
          s = s.strip()
          if s.startswith('```'):
            s = re.sub(r'^```json\s*', '', s, flags=re.IGNORECASE)
            s = re.sub(r'^```', '', s)
            s = re.sub(r'```\s*$', '', s)
          lb = s.find('[')
          rb = s.rfind(']')
          frag = s[lb:rb+1] if lb != -1 and rb != -1 and rb > lb else s
          frag = re.sub(r',\s*([\]\}])', r'\1', frag)
          items = json.loads(frag)
          if not isinstance(items, list):
            items = []
        except Exception as parse_e:
          debug['parse_error'] = str(parse_e)
          debug['raw_text_excerpt'] = str(text)[:500]
          print({'AI_PARSE_ERROR': debug})
          items = []
      else:
        try:
          debug['response_excerpt'] = r.text[:1000]
        except Exception:
          debug['response_excerpt'] = None
        debug['status_code'] = r.status_code
        print({'AI_UPSTREAM_NOT_200': debug})
        try:
          if 'error counting image token' in str(debug.get('response_excerpt') or ''):
            payload['stream'] = False
            r2 = requests.post(api_url, json=payload, headers={'Authorization': f'Bearer {key}', 'Content-Type': 'application/json'}, timeout=1200)
            if r2.status_code == 200:
              data2 = r2.json()
              text = data2.get('choices', [{}])[0].get('message', {}).get('content', '')
              try:
                import json, re
                s = str(text or '').strip()
                if s.startswith('```'):
                  s = re.sub(r'^```json\s*', '', s, flags=re.IGNORECASE)
                  s = re.sub(r'^```', '', s)
                  s = re.sub(r'```\s*$', '', s)
                lb = s.find('[')
                rb = s.rfind(']')
                frag = s[lb:rb+1] if lb != -1 and rb != -1 and rb > lb else s
                frag = re.sub(r',\s*([\]\}])', r'\1', frag)
                items = json.loads(frag)
                if not isinstance(items, list):
                  items = []
                return jsonify({'items': items})
              except Exception:
                pass
        except Exception:
          pass
        return jsonify({'error': 'upstream_status_not_200', 'debug': debug}), 502
    return jsonify({'items': items})
  except Exception as e:
    print({'AI_FATAL': str(e)})
    return jsonify({'error': str(e)}), 500

@app.get('/api/v1/ledger/download')
def ledger_download():
  try:
    p = (request.args.get('path') or '').strip()
    if not p:
      return jsonify({'error': '缺少 path'}), 400
    if not os.path.isfile(p):
      return jsonify({'error': '文件不存在'}), 404
    return send_file(p, as_attachment=True)
  except Exception as e:
    import traceback
    try:
      return jsonify({'error': str(e), 'stage': 'fatal', 'debug': dbg, 'trace': traceback.format_exc()}), 500
    except Exception:
      return jsonify({'error': str(e)}), 500

@app.get('/api/v1/openapi.json')
def openapi_spec():
  return jsonify({
    'openapi': '3.0.0',
    'info': {'title': 'Ledger API', 'version': '1.0.0'},
    'paths': {
      '/api/v1/ledger/process': {
        'post': {
          'summary': '生成台账',
          'requestBody': {'required': True},
          'responses': {'200': {'description': 'OK'}}
        }
      },
      '/api/v1/ledger/image-recognize': {
        'post': {
          'summary': '图片识别',
          'requestBody': {'required': True},
          'responses': {'200': {'description': 'OK'}}
        }
      },
      '/api/v1/ledger/external-upload': {
        'post': {
          'summary': '台账外部上传',
          'requestBody': {'required': True},
          'responses': {'200': {'description': 'OK'}}
        }
      },
      '/api/v1/ledger/external-upload/status': {
        'get': {
          'summary': '台账外部上传状态查询',
          'responses': {'200': {'description': 'OK'}}
        }
      }
    }
  })

@app.get('/api/v1/order/load')
def order_load_key():
  try:
    key = (request.args.get('key') or '').strip()
    if not key:
      return jsonify({'error': '缺少 key'}), 400
    store = _read_order_store()
    return jsonify({'data': store.get(key), 'lastModified': store.get('lastModified')})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.get('/api/v1/order/items-range')
def order_items_range():
  try:
    from datetime import datetime
    def _parse_date(s):
      try:
        return datetime.strptime(s, '%Y-%m-%d')
      except Exception:
        return None
    start = (request.args.get('start') or '').strip()
    end = (request.args.get('end') or '').strip()
    sd = _parse_date(start)
    ed = _parse_date(end)
    store = _read_order_store()
    items_by_date = store.get('items_by_date') or {}
    out = []
    for d, arr in items_by_date.items():
      dt = _parse_date(d)
      if dt is None:
        continue
      if sd and dt < sd:
        continue
      if ed and dt > ed:
        continue
      out.append({'date': d, 'items': arr})
    out.sort(key=lambda x: x['date'])
    return jsonify({'data': out})
  except Exception as e:
    return jsonify({'error': str(e)}), 500

# ===== Frontend Static Serving =====
FRONTEND_DIST = os.path.join(os.path.dirname(__file__), 'dist')

@app.route('/')
def index():
    return send_file(os.path.join(FRONTEND_DIST, 'index.html'))

@app.route('/<path:path>')
def serve_static(path):
    p = os.path.join(FRONTEND_DIST, path)
    if os.path.exists(p):
        return send_file(p)
    # Support client-side routing by falling back to index.html for non-API routes
    if path.startswith('api/') or path.startswith('uploads/'):
        return jsonify({'error': 'Not Found'}), 404
    return send_file(os.path.join(FRONTEND_DIST, 'index.html'))

@app.get('/api/v1/weekly-sales-summary')
def weekly_sales_summary():
  start_date = request.args.get('start_date')
  data = gb.get_weekly_sales_summary(start_date)
  return jsonify(data)
@app.get('/api/v1/staff/weekly-schedule')
def staff_weekly_schedule():
  name = (request.args.get('name') or '').strip()
  start_date = (request.args.get('start_date') or '').strip()
  if not start_date:
    start_date = datetime.today().strftime('%Y-%m-%d')
  try:
    base = datetime.strptime(start_date, '%Y-%m-%d')
  except Exception:
    return jsonify({'error': 'start_date 格式应为 YYYY-MM-DD'}), 400

  week = []
  for i in range(7):
    d = base + timedelta(days=i)
    ds = d.strftime('%Y-%m-%d')
    try:
      ban = gb.get_banbiao_data(ds) or {}
      staff_list = ban.get('staffList') or []
      entry = None
      for s in staff_list:
        # 支持模糊包含匹配，避免姓名空格差异
        sn = str(s.get('staffNm') or '')
        if not name or (sn and (name in sn or sn in name)):
          entry = s
          break
      item = {
        'date': ds,
        'weekday': ['星期一','星期二','星期三','星期四','星期五','星期六','星期日'][d.weekday()],
      }
      if entry:
        # 岗位与休息信息字段兼容多种命名
        position = entry.get('industry') or entry.get('position') or entry.get('岗位')
        break_start = entry.get('breakStart') or entry.get('restStart') or entry.get('休息开始')
        break_end = entry.get('breakEnd') or entry.get('restEnd') or entry.get('休息结束')
        break_time = None
        if break_start or break_end:
          bs = str(break_start or '')
          be = str(break_end or '')
          break_time = f"{bs}-{be}".strip('-')

        def _parse_hhmm(s):
          try:
            if s is None:
              return None
            s = str(s).replace(':','')
            if len(s) < 3:
              return None
            h = int(s[:2])
            m = int(s[2:4]) if len(s) >= 4 else 0
            return h*60 + m
          except Exception:
            return None

        def _diff_minutes(start, end):
          if start is None or end is None:
            return 0
          diff = end - start
          if diff < 0:
            diff += 24*60
          return diff

        ss = _parse_hhmm(entry.get('shiftStart'))
        se = _parse_hhmm(entry.get('shiftEnd'))
        bs_min = _parse_hhmm(break_start)
        be_min = _parse_hhmm(break_end)
        total_min = _diff_minutes(ss, se)
        break_min = _diff_minutes(bs_min, be_min)
        labor_hours = round(max(0, total_min - break_min) / 60, 2)
        break_hours = round(break_min / 60, 2)

        item.update({
          'name': entry.get('staffNm'),
          'shiftStart': entry.get('shiftStart'),
          'shiftEnd': entry.get('shiftEnd'),
          'laborTime': entry.get('laborTime'),
          'laborHours': labor_hours,
          'position': position,
          'breakTime': break_time,
          'breakHours': break_hours,
          'has_data': True,
        })
      else:
        item.update({'has_data': False})
      week.append(item)
    except Exception:
      week.append({'date': ds, 'weekday': ['星期一','星期二','星期三','星期四','星期五','星期六','星期日'][d.weekday()], 'has_data': False})

  return jsonify({'name': name, 'start_date': start_date, 'weekly_schedule': week})

@app.get('/api/v1/staff/daily-schedule')
def staff_daily_schedule():
  date = (request.args.get('date') or '').strip()
  if not date:
    return jsonify({'error': 'date 必填 YYYY-MM-DD'}), 400
  try:
    datetime.strptime(date, '%Y-%m-%d')
  except Exception:
    return jsonify({'error': 'date 格式应为 YYYY-MM-DD'}), 400
  try:
    ban = gb.get_banbiao_data(date) or {}
    staff_list = ban.get('staffList') or []
    sales_plan = ban.get('salesPlan') or {}
    out = []
    labor_total = 0.0
    for s in staff_list:
      position = s.get('industry') or s.get('position') or s.get('岗位')
      break_start = s.get('breakStart') or s.get('restStart') or s.get('休息开始')
      break_end = s.get('breakEnd') or s.get('restEnd') or s.get('休息结束')
      def _parse_hhmm(v):
        try:
          if v is None:
            return None
          v = str(v).replace(':','')
          if len(v) < 3:
            return None
          h = int(v[:2])
          m = int(v[2:4]) if len(v) >= 4 else 0
          return h*60 + m
        except Exception:
          return None
      def _diff(start, end):
        if start is None or end is None:
          return 0
        d = end - start
        if d < 0:
          d += 24*60
        return d
      ss = _parse_hhmm(s.get('shiftStart'))
      se = _parse_hhmm(s.get('shiftEnd'))
      bs = _parse_hhmm(break_start)
      be = _parse_hhmm(break_end)
      total = _diff(ss, se)
      brk = _diff(bs, be)
      labor_hours = round(max(0, total - brk) / 60, 2)
      break_hours = round(brk / 60, 2)
      labor_total += labor_hours
      break_time = None
      if break_start or break_end:
        bs_s = str(break_start or '')
        be_s = str(break_end or '')
        break_time = f"{bs_s}-{be_s}".strip('-')
      out.append({
        'date': date,
        'name': s.get('staffNm'),
        'position': position,
        'shiftStart': s.get('shiftStart'),
        'shiftEnd': s.get('shiftEnd'),
        'breakTime': break_time,
        'laborHours': labor_hours,
        'breakHours': break_hours,
      })
    # compute plan total and sales per labor hour
    def _to_float(v):
      try:
        return float(str(v).replace(',', ''))
      except Exception:
        return 0.0
    plan_total = 0.0
    if isinstance(sales_plan, dict):
      for _, val in sales_plan.items():
        plan_total += _to_float(val)
    splh = round(plan_total / labor_total, 2) if labor_total > 0 else 0.0
    return jsonify({
      'date': date,
      'staff': out,
      'salesPlan': sales_plan,
      'planTotal': round(plan_total, 2),
      'laborHoursTotal': round(labor_total, 2),
      'salesPerLaborHour': splh,
    })
  except Exception as e:
    return jsonify({'error': str(e)}), 500

@app.get('/api/v1/weekly-schedule')
def staff_weekly_schedule_alias():
  return staff_weekly_schedule()

@app.get('/api/v1/attendance')
def attendance():
  year_month = request.args.get('year_month')
  name = request.args.get('name')
  code = request.args.get('code')
  data = gb.get_attendance_data(year_month, name, code)
  return jsonify(data)

@app.get('/api/v1/revenue')
def revenue():
  try:
    data = gb.get_revenue()
  except Exception as e:
    print(f"Error fetching revenue: {e}")
    data = ""
  return jsonify({'revenue': data})

@app.get('/api/v1/hourly-sales')
def hourly_sales():
    from menu_system.db import SessionLocal
    from menu_system.models import HourlySalesSnapshot
    
    date_param = request.args.get('date')
    if not date_param:
        date_param = datetime.now().strftime('%Y-%m-%d')
    
    # Normalize date to YYYYMMDD
    try:
        if '-' in date_param:
            dt_obj = datetime.strptime(date_param, '%Y-%m-%d')
            date_str = dt_obj.strftime('%Y%m%d')
        else:
            date_str = date_param
    except:
        return jsonify({'error': 'Invalid date format'}), 400

    out = []
    try:
        with SessionLocal() as s:
            snapshots = s.query(HourlySalesSnapshot).filter_by(date=date_str).order_by(HourlySalesSnapshot.hour).all()
            
            # Calculate hourly delta
            prev_cumulative = 0.0
            # Sort by hour just in case
            snapshots.sort(key=lambda x: x.hour)
            
            for snap in snapshots:
                # If hour is 10 (start), previous is 0.
                # If we missed an hour (e.g. have 10 and 12), the delta for 12 will be (12 - 10), which is correct (sales during 10-12).
                # But we label it as "12".
                
                # Logic:
                # hour 10: sales up to 10.
                # hour 11: sales between 10 and 11.
                
                # However, if we missed 10, and first is 11. 
                # Then 11 is (sales up to 11).
                # So we just subtract the previous snapshot's cumulative.
                
                # Special handling: If this is the very first snapshot of the day, and it's later than 10 (e.g. 12),
                # we assume previous is 0. So 12 will contain all sales up to 12.
                
                delta = snap.cumulative_sales - prev_cumulative
                if delta < 0: delta = 0 # Should not happen unless refund > sales
                
                out.append({
                    'hour': snap.hour,
                    'sales': delta, 
                    'cumulative': snap.cumulative_sales,
                    'created_at': snap.created_at.isoformat() if snap.created_at else None
                })
                prev_cumulative = snap.cumulative_sales
                
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
    return jsonify({'date': date_str, 'data': out})


def _db_path() -> str:
  return os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'product', 'data.sqlite'))

def _fetch_borrow_records(date: str | None) -> List[Dict[str, Any]]:
  conn = sqlite3.connect(_db_path())
  conn.row_factory = sqlite3.Row
  try:
    cur = conn.cursor()
    if date:
      cur.execute('SELECT * FROM records WHERE borrowDate = ? ORDER BY id DESC', (date,))
    else:
      cur.execute('SELECT * FROM records ORDER BY id DESC')
    records = [dict(r) for r in cur.fetchall()]
    items_stmt = conn.cursor()
    out = []
    for r in records:
      items_stmt.execute('SELECT * FROM items WHERE recordId = ? ORDER BY id ASC', (r['id'],))
      items = [dict(i) for i in items_stmt.fetchall()]
      out.append({**r, 'items': items})
    return out
  finally:
    conn.close()

@app.get('/api/v1/borrow/records')
def borrow_records():
  date = request.args.get('date')
  data = _fetch_borrow_records(date)
  return jsonify(data)

def _conn():
  p = _db_path()
  conn = sqlite3.connect(p)
  conn.row_factory = sqlite3.Row
  conn.execute('PRAGMA foreign_keys = ON')
  return conn

def _compute_status(items: List[Dict[str, Any]]) -> str:
  if not items:
    return 'pending'
  all_ret = all((int(i.get('returnedQuantity') or 0)) >= int(i.get('quantity') or 0) for i in items)
  any_ret = any(0 < int(i.get('returnedQuantity') or 0) < int(i.get('quantity') or 0) for i in items)
  if all_ret:
    return 'returned'
  if any_ret:
    return 'partial'
  return 'pending'

def _record_with_items(conn, rid: int) -> Dict[str, Any] | None:
  r = conn.execute('SELECT * FROM records WHERE id = ?', (rid,)).fetchone()
  if not r:
    return None
  items = conn.execute('SELECT * FROM items WHERE recordId = ? ORDER BY id ASC', (rid,)).fetchall()
  return {**dict(r), 'items': [dict(i) for i in items]}

@app.post('/api/v1/borrow/records')
def borrow_create():
  payload = request.json or {}
  borrower = payload.get('borrower')
  borrowDate = payload.get('borrowDate')
  borrowUnit = payload.get('borrowUnit')
  sourceUnit = payload.get('sourceUnit')
  sourcePerson = payload.get('sourcePerson')
  notes = payload.get('notes') or ''
  items = payload.get('items') or []
  if not borrower or not borrowDate or not borrowUnit or not sourceUnit or not sourcePerson:
    return jsonify({'error': '缺少必要字段'}), 400
  if not isinstance(items, list) or len(items) == 0:
    return jsonify({'error': '至少需要一件商品'}), 400
  conn = _conn()
  try:
    status = _compute_status(items)
    cur = conn.cursor()
    cur.execute(
      'INSERT INTO records (borrower, borrowDate, borrowUnit, sourceUnit, sourcePerson, notes, status) VALUES (?,?,?,?,?,?,?)',
      (borrower, borrowDate, borrowUnit, sourceUnit, sourcePerson, notes, status)
    )
    rid = cur.lastrowid
    ins = conn.cursor()
    for it in items:
      name = it.get('name')
      if not name:
        continue
      spec = it.get('spec') or ''
      qty = int(it.get('quantity') or 0)
      ret = int(it.get('returnedQuantity') or 0)
      ins.execute('INSERT INTO items (recordId, name, spec, quantity, returnedQuantity) VALUES (?,?,?,?,?)', (rid, name, spec, qty, ret))
    conn.commit()
    out = _record_with_items(conn, rid)
    return jsonify(out), 201
  except Exception as e:
    return jsonify({'error': str(e)}), 500
  finally:
    conn.close()

@app.delete('/api/v1/borrow/records/<int:rid>')
def borrow_delete(rid: int):
  conn = _conn()
  try:
    conn.execute('DELETE FROM items WHERE recordId = ?', (rid,))
    conn.execute('DELETE FROM records WHERE id = ?', (rid,))
    conn.commit()
    return jsonify({'ok': True})
  except Exception as e:
    return jsonify({'error': str(e)}), 500
  finally:
    conn.close()

@app.post('/api/v1/borrow/records/<int:rid>/return')
def borrow_return(rid: int):
  payload = request.json or {}
  returns = payload.get('returns') or []
  returnDate = payload.get('returnDate')
  returnNotes = payload.get('returnNotes') or ''
  if not isinstance(returns, list):
    return jsonify({'error': 'returns 必须为数组'}), 400
  conn = _conn()
  try:
    items = [dict(i) for i in conn.execute('SELECT * FROM items WHERE recordId = ?', (rid,)).fetchall()]
    for r in returns:
      itemId = int(r.get('itemId') or 0)
      qty = int(r.get('qty') or 0)
      it = next((i for i in items if int(i['id']) == itemId), None)
      if not it:
        continue
      remaining = max(0, int(it['quantity']) - int(it['returnedQuantity']))
      applyQty = min(remaining, max(0, qty))
      newReturned = int(it['returnedQuantity']) + applyQty
      conn.execute('UPDATE items SET returnedQuantity = ? WHERE id = ?', (newReturned, itemId))
    updated = [dict(i) for i in conn.execute('SELECT * FROM items WHERE recordId = ?', (rid,)).fetchall()]
    status = _compute_status(updated)
    conn.execute('UPDATE records SET status = ?, returnDate = ?, returnNotes = ? WHERE id = ?', (status, returnDate, returnNotes, rid))
    conn.commit()
    out = _record_with_items(conn, rid)
    return jsonify(out)
  except Exception as e:
    return jsonify({'error': str(e)}), 500
  finally:
    conn.close()

@app.get('/api/v1/employees')
def get_employees_list():
  """
  获取员工列表
  ---
  tags:
    - Employee
  responses:
    200:
      description: 成功获取员工列表
      content:
        application/json:
          schema:
            type: object
            properties:
              code:
                type: integer
                example: 200
              message:
                type: string
                example: success
              data:
                type: array
                items:
                  type: object
                  properties:
                    id:
                      type: string
                      description: 员工ID
                    name:
                      type: string
                      description: 员工姓名
  """
  try:
    data = gb.get_employees()
    return jsonify({
      "code": 200,
      "message": "success",
      "data": data
    })
  except Exception as e:
    return jsonify({
      "code": 500,
      "message": str(e),
      "data": []
    }), 500

@app.get('/api/v1/voice-reminders')
def voice_reminders_get():
    from menu_system.services import get_voice_reminders
    reminders = get_voice_reminders()
    return jsonify([
        {
            'id': r.id,
            'time': r.time,
            'content': r.content,
            'enabled': r.enabled,
            'reminder_type': getattr(r, 'reminder_type', 'ai_voice'),
            'voice_model': getattr(r, 'voice_model', 'tts-1'),
            'audio_file_path': getattr(r, 'audio_file_path', None)
        }
        for r in reminders
    ])

@app.get('/api/v1/tts-models')
def get_tts_models():
    """获取可用的 TTS 模型列表"""
    models = [
        {
            'id': 'tts-1',
            'name': 'TTS-1 (标准)',
            'description': '速度快，成本低',
            'speed': '快',
            'quality': '标准'
        },
        {
            'id': 'tts-1-hd',
            'name': 'TTS-1-HD (高清)',
            'description': '音质更好，速度稍慢',
            'speed': '慢',
            'quality': '高清'
        }
    ]
    return jsonify({'models': models})

@app.post('/api/v1/voice-reminders')
def voice_reminders_create():
    from menu_system.services import create_voice_reminder
    body = request.json or {}
    time = body.get('time')
    content = body.get('content')
    if not time or not content:
        return jsonify({'error': 'Missing time or content'}), 400
    
    r = create_voice_reminder(
        time=time,
        content=content,
        reminder_type=body.get('reminder_type', 'ai_voice'),
        voice_model=body.get('voice_model', 'tts-1'),
        audio_file_path=body.get('audio_file_path')
    )
    
    return jsonify({
        'id': r.id,
        'time': r.time,
        'content': r.content,
        'enabled': r.enabled,
        'reminder_type': getattr(r, 'reminder_type', 'ai_voice'),
        'voice_model': getattr(r, 'voice_model', 'tts-1'),
        'audio_file_path': getattr(r, 'audio_file_path', None)
    }), 201

@app.post('/api/v1/voice-reminders/batch')
def voice_reminders_batch_create():
    from menu_system.services import batch_create_voice_reminders
    body = request.json or {}
    items = body.get('items')
    if not items or not isinstance(items, list):
        return jsonify({'error': 'Missing items or items is not a list'}), 400
    results = batch_create_voice_reminders(items)
    return jsonify([
        {'id': r.id, 'time': r.time, 'content': r.content, 'enabled': r.enabled}
        for r in results
    ]), 201

@app.patch('/api/v1/voice-reminders/<int:rid>')
def voice_reminders_update(rid):
    from menu_system.services import update_voice_reminder
    body = request.json or {}
    r = update_voice_reminder(
        rid,
        time=body.get('time'),
        content=body.get('content'),
        enabled=body.get('enabled'),
        reminder_type=body.get('reminder_type'),
        voice_model=body.get('voice_model'),
        audio_file_path=body.get('audio_file_path')
    )
    if not r:
        return jsonify({'error': 'Not found'}), 404
    return jsonify({
        'id': r.id,
        'time': r.time,
        'content': r.content,
        'enabled': r.enabled,
        'reminder_type': getattr(r, 'reminder_type', 'ai_voice'),
        'voice_model': getattr(r, 'voice_model', 'tts-1'),
        'audio_file_path': getattr(r, 'audio_file_path', None)
    })

@app.delete('/api/v1/voice-reminders/<int:rid>')
def voice_reminders_delete(rid):
    from menu_system.services import delete_voice_reminder
    if delete_voice_reminder(rid):
        return jsonify({'ok': True})
    return jsonify({'error': 'Not found'}), 404

@app.delete('/api/v1/voice-reminders')
def voice_reminders_delete_all():
    from menu_system.services import delete_all_voice_reminders
    delete_all_voice_reminders()
    return jsonify({'ok': True})

@app.post('/api/v1/voice-reminders/upload-audio')
def upload_reminder_audio():
    """上传自定义音频文件"""
    import os
    from werkzeug.utils import secure_filename
    import time as time_module
    
    UPLOAD_FOLDER = 'data/uploads/audio'
    ALLOWED_EXTENSIONS = {'mp3', 'wav', 'ogg', 'm4a', 'aac'}
    
    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        # 确保上传目录存在
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
        # 生成唯一文件名
        timestamp = int(time_module.time() * 1000)
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"reminder_{timestamp}.{ext}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        # 保存文件
        file.save(filepath)
        
        # 返回相对路径
        relative_path = f"/uploads/audio/{filename}"
        return jsonify({
            'file_path': relative_path,
            'filename': filename
        })
    
    return jsonify({'error': 'Invalid file type. Allowed: mp3, wav, ogg, m4a, aac'}), 400

@app.get('/uploads/audio/<filename>')
def serve_audio(filename):
    """提供音频文件访问"""
    from flask import send_from_directory
    import os
    UPLOAD_FOLDER = 'data/uploads/audio'
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.post('/api/v1/push-subscriptions')
def push_subscriptions_add():
    from menu_system.services import add_push_subscription
    body = request.json or {}
    endpoint = body.get('endpoint')
    keys = body.get('keys') or {}
    p256dh = keys.get('p256dh')
    auth = keys.get('auth')
    if not endpoint or not p256dh or not auth:
        return jsonify({'error': 'Missing push data'}), 400
    add_push_subscription(endpoint, p256dh, auth)
    return jsonify({'ok': True})

@app.delete('/api/v1/push-subscriptions')
def push_subscriptions_delete():
    from menu_system.services import delete_push_subscription
    body = request.json or {}
    endpoint = body.get('endpoint')
    if not endpoint:
        return jsonify({'error': 'Missing endpoint'}), 400
    delete_push_subscription(endpoint)
    return jsonify({'ok': True})

def get_or_create_vapid_keys():
    vapid_file = os.path.join(UPLOAD_DIR, 'vapid_keys.json')
    private_pem_file = os.path.join(UPLOAD_DIR, 'vapid_private.pem')
    
    if os.path.exists(vapid_file) and os.path.exists(private_pem_file):
        try:
            with open(vapid_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                data['private_key_path'] = private_pem_file
                return data
        except:
            pass
            
    # Generate P-256 key pair
    private_key = ec.generate_private_key(ec.SECP256R1())
    
    # Export Private Key to PEM (Traditional OpenSSL SEC1 format)
    private_pem = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.TraditionalOpenSSL,
        encryption_algorithm=serialization.NoEncryption()
    )
    
    _ensure_dir(UPLOAD_DIR)
    with open(private_pem_file, 'wb') as f:
        f.write(private_pem)
        
    # Export Public Key for frontend
    public_key = private_key.public_key()
    public_bytes = public_key.public_bytes(
        encoding=serialization.Encoding.X962,
        format=serialization.PublicFormat.UncompressedPoint
    )
    public_b64url = base64.urlsafe_b64encode(public_bytes).decode('utf-8').rstrip('=')
    
    v_data = {
        'public_key': public_b64url,
        'private_key': private_pem.decode('utf-8') # Keep for compatibility if needed
    }
    
    with open(vapid_file, 'w', encoding='utf-8') as f:
        json.dump(v_data, f)
    
    v_data['private_key_path'] = private_pem_file
    return v_data

@app.get('/api/v1/vapid-public-key')
def vapid_public_key():
    v_data = get_or_create_vapid_keys()
    if v_data:
        return jsonify({'publicKey': v_data['public_key']})
    return jsonify({'error': 'Failed to handle VAPID keys'}), 500

@app.delete('/api/v1/vapid-keys')
def reset_vapid_keys():
    vapid_file = os.path.join(UPLOAD_DIR, 'vapid_keys.json')
    private_pem_file = os.path.join(UPLOAD_DIR, 'vapid_private.pem')
    if os.path.exists(vapid_file):
        os.remove(vapid_file)
    if os.path.exists(private_pem_file):
        os.remove(private_pem_file)
    return jsonify({'ok': True})

def log_push(endpoint, status, message=""):
    try:
        log_file = os.path.join(UPLOAD_DIR, 'push_logs.json')
        _ensure_dir(UPLOAD_DIR)
        
        logs = []
        if os.path.exists(log_file):
            with open(log_file, 'r', encoding='utf-8') as f:
                content = f.read()
                if content:
                    logs = json.loads(content)
        
        logs.append({
            'time': datetime.now().strftime('%m-%d %H:%M:%S'),
            'endpoint': str(endpoint)[:30] + "...",
            'status': str(status),
            'message': str(message)
        })
        
        if len(logs) > 30:
            logs = logs[-30:]
            
        with open(log_file, 'w', encoding='utf-8') as f:
            json.dump(logs, f, ensure_ascii=False)
    except Exception as e:
        print(f"CRITICAL: log_push failed: {e}")

def send_web_push(subscription, data):
    from pywebpush import webpush, WebPushException
    
    print(f"DEBUG: Attempting push to {subscription.endpoint[:30]}...")
    log_push(subscription.endpoint, "Attempt", "Starting push process")
    
    v_data = get_or_create_vapid_keys()
    if not v_data or 'private_key' not in v_data:
        log_push(subscription.endpoint, "Error", "VAPID keys missing")
        return
        
    try:
        from urllib.parse import urlparse
        parsed_url = urlparse(subscription.endpoint)
        audience = f"{parsed_url.scheme}://{parsed_url.netloc}"
        
        webpush(
            subscription_info={
                'endpoint': subscription.endpoint,
                'keys': {
                    'p256dh': subscription.p256dh,
                    'auth': subscription.auth
                }
            },
            data=json.dumps(data),
            # Use the absolute path to the PEM file for maximum reliability
            vapid_private_key=os.path.abspath(v_data['private_key_path']),
            vapid_claims={
                "sub": "mailto:admin@muhuo.site",
                "aud": audience
            },
            ttl=86400
        )
        log_push(subscription.endpoint, "Success")
    except WebPushException as ex:
        err_msg = f"WebPush Error: {ex}"
        if ex.response is not None:
            err_msg = f"HTTP {ex.response.status_code}: {ex.response.text}"
            print(f"DEBUG: Push Failed with response: {ex.response.text}")
            
            # Special handling for VapidPkHashMismatch - this means the client MUST re-subscribe
            if "VapidPkHashMismatch" in ex.response.text:
                err_msg = "VAPID Key Mismatch: Please click 'Reset System' on your phone"
                
        log_push(subscription.endpoint, "Fail", err_msg)
        if ex.response and (ex.response.status_code in [404, 410] or "VapidPkHashMismatch" in ex.response.text):
            from menu_system.services import delete_push_subscription
            delete_push_subscription(subscription.endpoint)
            print(f"DEBUG: Deleted invalid subscription: {subscription.endpoint[:30]}...")
    except Exception as e:
        log_push(subscription.endpoint, "Error", str(e))

@app.get('/api/v1/push-logs')
def get_push_logs():
    log_file = os.path.join(UPLOAD_DIR, 'push_logs.json')
    if os.path.exists(log_file):
        try:
            with open(log_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return jsonify(data if isinstance(data, list) else [])
        except Exception as e:
            print(f"Error reading push logs: {e}")
            return jsonify([{'time': 'Error', 'status': 'Error', 'message': str(e)}])
    return jsonify([])

@app.post('/api/v1/test-push')
def test_push_now():
    from menu_system.db import SessionLocal
    from menu_system.models import PushSubscription
    with SessionLocal() as s:
        subs = s.query(PushSubscription).all()
        if not subs:
            return jsonify({'error': 'No subscriptions found'}), 404
        for sub in subs:
            send_web_push(sub, {
                'title': '推送测试',
                'body': '看到这条消息说明推送配置正确！',
                'type': 'test'
            })
    return jsonify({'ok': True, 'count': len(subs)})

def reminder_scheduler_task():
    from menu_system.db import SessionLocal
    from menu_system.models import VoiceReminder, PushSubscription
    import time
    
    print("Reminder scheduler task started.")
    last_checked_minute = ""
    
    while True:
        try:
            # Use CST (China Standard Time) or local time properly
            # iOS users are in China, so we assume UTC+8 if server is UTC
            # Or just use the server's system time if it's already set to China time
            now = datetime.now()
            
            # If server is in UTC, add 8 hours to match user's China time
            # Check if current hour is logically consistent with restaurant hours
            # If server time is 0-14, it might be UTC. China 8-22 is UTC 0-14.
            # This is a heuristic, better to use pytz if available.
            # Let's try to detect if server is UTC
            if time.tzname[0] == 'UTC':
                now = now + timedelta(hours=8)
                
            current_hm = now.strftime('%H:%M')
            
            if current_hm != last_checked_minute:
                last_checked_minute = current_hm
                
                with SessionLocal() as s:
                    # Find enabled reminders for this minute
                    reminders = s.query(VoiceReminder).filter_by(time=current_hm, enabled=True).all()
                    
                    if reminders:
                        subscriptions = s.query(PushSubscription).all()
                        for r in reminders:
                            print(f"Pushing reminder: {r.content}")
                            for sub in subscriptions:
                                # Send push notification
                                send_web_push(sub, {
                                    'title': '食材过期提醒',
                                    'body': r.content,
                                    'type': 'voice_reminder'
                                })
            
            time.sleep(10)
        except Exception as e:
            print(f"Reminder scheduler error: {e}")
            time.sleep(10)

def revenue_scheduler_task():
  from menu_system.db import SessionLocal
  from menu_system.models import RevenueData, HourlySalesSnapshot
  import time
  
  print("Revenue scheduler task started.")
  while True:
    try:
      now = datetime.now()
      # Fetch between 10:00 and 22:00
      if 10 <= now.hour <= 22:
        should_run = False
        today_str = now.strftime('%Y%m%d')
        current_hour = now.hour
        
        # 1. Update RevenueData (Snapshot for /api/v1/revenue) - Keep existing logic to update every 50 mins or so
        # Actually, let's just piggyback on the hourly schedule for simplicity, OR keep it frequent?
        # User requirement: "Automatically trigger once every hour on the hour"
        # So we should strictly follow that for the HourlySalesSnapshot.
        
        # Check if we have snapshot for this hour
        has_snapshot = False
        try:
          with SessionLocal() as s:
             snap = s.query(HourlySalesSnapshot).filter_by(date=today_str, hour=current_hour).first()
             if snap:
               has_snapshot = True
        except Exception as e:
           print(f"Scheduler DB check error: {e}")
        
        # Run if it's the start of the hour (minute < 5) and we don't have a snapshot yet
        # Retry mechanism is implicit: if we fail, we wait 60s and try again (minute still < 5)
        # If we miss the 5 min window, we miss the data.
        if now.minute < 5 and not has_snapshot:
             should_run = True
        
        # Also run if it's 10:00 and we have no data at all? 
        # No, just stick to the schedule.

        if should_run:
          print(f"Scheduler fetching revenue data at {now}")
          data_str = gb.get_revenue()
          
          # Parse cumulative sales
          # data_str is like "12345" or "12345, 67890"
          cumulative_val = 0.0
          try:
             parts = str(data_str).split(',')
             val_str = parts[0].strip()
             cumulative_val = float(val_str)
          except:
             cumulative_val = 0.0

          with SessionLocal() as s:
            # 1. Save Hourly Snapshot
            # Double check inside transaction
            existing = s.query(HourlySalesSnapshot).filter_by(date=today_str, hour=current_hour).first()
            if not existing:
                new_snap = HourlySalesSnapshot(
                    date=today_str, 
                    hour=current_hour, 
                    cumulative_sales=cumulative_val,
                    created_at=datetime.utcnow()
                )
                s.add(new_snap)
            
            # 2. Update Latest RevenueData (for legacy support)
            rd = s.query(RevenueData).filter(RevenueData.date == today_str).first()
            if not rd:
              rd = RevenueData(date=today_str)
              s.add(rd)
            rd.raw_data = data_str
            rd.updated_at = datetime.utcnow()
            
            s.commit()
            print(f"Revenue data saved for {today_str} {current_hour}:00")
      
      # Check every minute to ensure we catch the "minute < 5" window
      time.sleep(60)
    except Exception as e:
      print(f"Scheduler error: {e}")
      time.sleep(60)

if __name__ == '__main__':
  t1 = threading.Thread(target=revenue_scheduler_task, daemon=True)
  t1.start()
  t2 = threading.Thread(target=reminder_scheduler_task, daemon=True)
  t2.start()
  app.run(host='0.0.0.0', port=8000)
